"""
Data export/import routes for managing assignment data via Excel.
Admin-only access for data management.
"""
from fastapi import APIRouter, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.database import get_db, init_database
from app.dependencies import get_current_user, is_admin
from app.templates_config import templates

router = APIRouter()


def require_admin_access(request: Request):
    """Check if user is admin, return user or redirect."""
    user = get_current_user(request)
    if not user:
        return None, RedirectResponse(url="/login", status_code=302)
    if not is_admin(user):
        return None, RedirectResponse(url="/dashboard", status_code=302)
    return user, None


def get_config_options(category: str, parent_value: str = None) -> list:
    """Get configuration options from database."""
    with get_db() as conn:
        cursor = conn.cursor()
        if parent_value:
            cursor.execute("""
                SELECT option_value, option_label FROM config_options
                WHERE category = ? AND parent_value = ? AND is_active = 1
                ORDER BY sort_order
            """, (category, parent_value))
        else:
            cursor.execute("""
                SELECT option_value, option_label FROM config_options
                WHERE category = ? AND (parent_value IS NULL OR parent_value = '') AND is_active = 1
                ORDER BY sort_order
            """, (category,))
        return [dict(row) for row in cursor.fetchall()]


@router.get("/export", response_class=HTMLResponse)
async def export_page(request: Request):
    """Display data export options page. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    return templates.TemplateResponse("data_export.html", {
        "request": request,
        "user": user
    })


@router.get("/export/assignments")
async def export_assignments(request: Request):
    """Export all assignments to Excel. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    with get_db() as conn:
        cursor = conn.cursor()

        # Get all assignments with related data
        cursor.execute("""
            SELECT
                a.id, a.assignment_no, a.type, a.title, a.client, a.client_type,
                a.city, a.domain, a.sub_domain, a.office_id, a.status,
                a.tor_scope, a.work_order_date, a.start_date, a.target_date,
                a.team_leader_officer_id, a.venue, a.duration_start, a.duration_end,
                a.duration_days, a.type_of_participants,
                a.faculty1_officer_id, a.faculty2_officer_id,
                a.total_value, a.invoice_amount, a.amount_received,
                a.total_expenditure, a.total_revenue, a.surplus_deficit,
                a.physical_progress_percent, a.timeline_progress_percent,
                a.remarks
            FROM assignments a
            ORDER BY a.office_id, a.assignment_no
        """)
        assignments = cursor.fetchall()

        # Create workbook
        wb = Workbook()

        # Sheet 1: Assignments
        ws_assignments = wb.active
        ws_assignments.title = "Assignments"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            'ID', 'Assignment No', 'Type', 'Title', 'Client', 'Client Type',
            'City', 'Domain', 'Sub-Domain', 'Office', 'Status',
            'ToR/Scope', 'Work Order Date', 'Start Date', 'Target Date',
            'Team Leader ID', 'Venue', 'Duration Start', 'Duration End',
            'Duration Days', 'Type of Participants',
            'Faculty 1 ID', 'Faculty 2 ID',
            'Total Value (L)', 'Invoice Raised (L)', 'Payment Received (L)',
            'Total Expenditure (L)', 'Revenue Shared (L)', 'Surplus/Deficit (L)',
            'Physical Progress %', 'Timeline Progress %', 'Remarks'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_assignments.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row_idx, assignment in enumerate(assignments, 2):
            for col_idx, value in enumerate(assignment, 1):
                cell = ws_assignments.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in ws_assignments.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_assignments.column_dimensions[column].width = adjusted_width

        # Sheet 2: Milestones
        ws_milestones = wb.create_sheet("Milestones")
        milestone_headers = [
            'Assignment ID', 'Milestone No', 'Title', 'Description', 'Target Date',
            'Actual Completion Date', 'Revenue %', 'Revenue Amount (L)',
            'Invoice Raised', 'Invoice Raised Date',
            'Payment Received', 'Payment Received Date', 'Status', 'Remarks'
        ]

        for col, header in enumerate(milestone_headers, 1):
            cell = ws_milestones.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        cursor.execute("""
            SELECT assignment_id, milestone_no, title, description, target_date,
                   actual_completion_date, revenue_percent, revenue_amount,
                   invoice_raised, invoice_raised_date,
                   payment_received, payment_received_date, status, remarks
            FROM milestones
            ORDER BY assignment_id, milestone_no
        """)
        milestones = cursor.fetchall()

        for row_idx, milestone in enumerate(milestones, 2):
            for col_idx, value in enumerate(milestone, 1):
                cell = ws_milestones.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Sheet 3: Revenue Shares
        ws_revenue = wb.create_sheet("Revenue Shares")
        revenue_headers = ['Assignment ID', 'Officer ID', 'Share %', 'Share Amount (L)']

        for col, header in enumerate(revenue_headers, 1):
            cell = ws_revenue.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        cursor.execute("""
            SELECT assignment_id, officer_id, share_percent, share_amount
            FROM revenue_shares
            ORDER BY assignment_id, officer_id
        """)
        shares = cursor.fetchall()

        for row_idx, share in enumerate(shares, 2):
            for col_idx, value in enumerate(share, 1):
                cell = ws_revenue.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Sheet 4: Expenditure
        ws_expenditure = wb.create_sheet("Expenditure")
        exp_headers = ['Assignment ID', 'Head Code', 'Head Name', 'Estimated (L)', 'Actual (L)', 'Remarks']

        for col, header in enumerate(exp_headers, 1):
            cell = ws_expenditure.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        cursor.execute("""
            SELECT e.assignment_id, h.head_code, h.head_name,
                   e.estimated_amount, e.actual_amount, e.remarks
            FROM expenditure_items e
            JOIN expenditure_heads h ON e.head_id = h.id
            ORDER BY e.assignment_id, h.head_code
        """)
        expenditures = cursor.fetchall()

        for row_idx, exp in enumerate(expenditures, 2):
            for col_idx, value in enumerate(exp, 1):
                cell = ws_expenditure.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Sheet 5: Officers (Reference)
        ws_officers = wb.create_sheet("Officers (Ref)")
        officer_headers = ['Officer ID', 'Name', 'Designation', 'Office', 'Email', 'Annual Target (L)']

        for col, header in enumerate(officer_headers, 1):
            cell = ws_officers.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        cursor.execute("""
            SELECT officer_id, name, designation, office_id, email, annual_target
            FROM officers WHERE is_active = 1
            ORDER BY office_id, name
        """)
        officers = cursor.fetchall()

        for row_idx, officer in enumerate(officers, 2):
            for col_idx, value in enumerate(officer, 1):
                cell = ws_officers.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Sheet 6: Configuration Options
        ws_config = wb.create_sheet("Config Options")
        config_headers = ['Category', 'Value', 'Label', 'Parent Value', 'Sort Order', 'Active']

        for col, header in enumerate(config_headers, 1):
            cell = ws_config.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        cursor.execute("""
            SELECT category, option_value, option_label, parent_value, sort_order, is_active
            FROM config_options
            ORDER BY category, sort_order
        """)
        configs = cursor.fetchall()

        for row_idx, config in enumerate(configs, 2):
            for col_idx, value in enumerate(config, 1):
                cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PMS_Data_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/import", response_class=HTMLResponse)
async def import_page(request: Request):
    """Display data import page. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    return templates.TemplateResponse("data_import.html", {
        "request": request,
        "user": user
    })


@router.post("/import")
async def import_data(request: Request, file: UploadFile = File(...), import_mode: str = Form("update")):
    """Import data from Excel file. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    errors = []
    success_counts = {
        'assignments': 0,
        'milestones': 0,
        'revenue_shares': 0,
        'expenditure': 0,
        'config': 0
    }

    try:
        contents = await file.read()
        excel_file = io.BytesIO(contents)

        # Read all sheets
        xl = pd.ExcelFile(excel_file)

        with get_db() as conn:
            cursor = conn.cursor()

            # Import Assignments
            if 'Assignments' in xl.sheet_names:
                df = pd.read_excel(xl, 'Assignments')
                df = df.where(pd.notnull(df), None)

                for idx, row in df.iterrows():
                    try:
                        assignment_id = row.get('ID')
                        if pd.isna(assignment_id):
                            # Insert new assignment
                            cursor.execute("""
                                INSERT INTO assignments (
                                    assignment_no, type, title, client, client_type,
                                    city, domain, sub_domain, office_id, status,
                                    tor_scope, work_order_date, start_date, target_date,
                                    team_leader_officer_id, venue, duration_start, duration_end,
                                    duration_days, type_of_participants,
                                    faculty1_officer_id, faculty2_officer_id,
                                    total_value, invoice_amount, amount_received,
                                    total_expenditure, total_revenue, surplus_deficit,
                                    physical_progress_percent, timeline_progress_percent, remarks
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                row.get('Assignment No'), row.get('Type'), row.get('Title'),
                                row.get('Client'), row.get('Client Type'), row.get('City'),
                                row.get('Domain'), row.get('Sub-Domain'), row.get('Office'),
                                row.get('Status'), row.get('ToR/Scope'),
                                row.get('Work Order Date'), row.get('Start Date'), row.get('Target Date'),
                                row.get('Team Leader ID'), row.get('Venue'),
                                row.get('Duration Start'), row.get('Duration End'),
                                row.get('Duration Days'), row.get('Type of Participants'),
                                row.get('Faculty 1 ID'), row.get('Faculty 2 ID'),
                                row.get('Total Value (L)'), row.get('Invoice Raised (L)'),
                                row.get('Payment Received (L)'), row.get('Total Expenditure (L)'),
                                row.get('Revenue Shared (L)'), row.get('Surplus/Deficit (L)'),
                                row.get('Physical Progress %'), row.get('Timeline Progress %'),
                                row.get('Remarks')
                            ))
                        else:
                            # Update existing assignment
                            cursor.execute("""
                                UPDATE assignments SET
                                    assignment_no = ?, type = ?, title = ?, client = ?, client_type = ?,
                                    city = ?, domain = ?, sub_domain = ?, office_id = ?, status = ?,
                                    tor_scope = ?, work_order_date = ?, start_date = ?, target_date = ?,
                                    team_leader_officer_id = ?, venue = ?, duration_start = ?, duration_end = ?,
                                    duration_days = ?, type_of_participants = ?,
                                    faculty1_officer_id = ?, faculty2_officer_id = ?,
                                    total_value = ?, invoice_amount = ?, amount_received = ?,
                                    total_expenditure = ?, total_revenue = ?, surplus_deficit = ?,
                                    physical_progress_percent = ?, timeline_progress_percent = ?,
                                    remarks = ?, updated_at = CURRENT_TIMESTAMP
                                WHERE id = ?
                            """, (
                                row.get('Assignment No'), row.get('Type'), row.get('Title'),
                                row.get('Client'), row.get('Client Type'), row.get('City'),
                                row.get('Domain'), row.get('Sub-Domain'), row.get('Office'),
                                row.get('Status'), row.get('ToR/Scope'),
                                row.get('Work Order Date'), row.get('Start Date'), row.get('Target Date'),
                                row.get('Team Leader ID'), row.get('Venue'),
                                row.get('Duration Start'), row.get('Duration End'),
                                row.get('Duration Days'), row.get('Type of Participants'),
                                row.get('Faculty 1 ID'), row.get('Faculty 2 ID'),
                                row.get('Total Value (L)'), row.get('Invoice Raised (L)'),
                                row.get('Payment Received (L)'), row.get('Total Expenditure (L)'),
                                row.get('Revenue Shared (L)'), row.get('Surplus/Deficit (L)'),
                                row.get('Physical Progress %'), row.get('Timeline Progress %'),
                                row.get('Remarks'), int(assignment_id)
                            ))
                        success_counts['assignments'] += 1
                    except Exception as e:
                        errors.append(f"Assignment row {idx + 2}: {str(e)}")

            # Import Milestones
            if 'Milestones' in xl.sheet_names:
                df = pd.read_excel(xl, 'Milestones')
                df = df.where(pd.notnull(df), None)

                for idx, row in df.iterrows():
                    try:
                        cursor.execute("""
                            INSERT OR REPLACE INTO milestones (
                                assignment_id, milestone_no, title, description, target_date,
                                actual_completion_date, revenue_percent, revenue_amount,
                                invoice_raised, invoice_raised_date,
                                payment_received, payment_received_date, status, remarks
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            row.get('Assignment ID'), row.get('Milestone No'),
                            row.get('Title'), row.get('Description'), row.get('Target Date'),
                            row.get('Actual Completion Date'), row.get('Revenue %'),
                            row.get('Revenue Amount (L)'),
                            1 if row.get('Invoice Raised') in [1, '1', 'Yes', 'yes', True] else 0,
                            row.get('Invoice Raised Date'),
                            1 if row.get('Payment Received') in [1, '1', 'Yes', 'yes', True] else 0,
                            row.get('Payment Received Date'),
                            row.get('Status'), row.get('Remarks')
                        ))
                        success_counts['milestones'] += 1
                    except Exception as e:
                        errors.append(f"Milestone row {idx + 2}: {str(e)}")

            # Import Revenue Shares
            if 'Revenue Shares' in xl.sheet_names:
                df = pd.read_excel(xl, 'Revenue Shares')
                df = df.where(pd.notnull(df), None)

                for idx, row in df.iterrows():
                    try:
                        cursor.execute("""
                            INSERT OR REPLACE INTO revenue_shares (
                                assignment_id, officer_id, share_percent, share_amount
                            ) VALUES (?, ?, ?, ?)
                        """, (
                            row.get('Assignment ID'), row.get('Officer ID'),
                            row.get('Share %'), row.get('Share Amount (L)')
                        ))
                        success_counts['revenue_shares'] += 1
                    except Exception as e:
                        errors.append(f"Revenue Share row {idx + 2}: {str(e)}")

            # Import Expenditure
            if 'Expenditure' in xl.sheet_names:
                df = pd.read_excel(xl, 'Expenditure')
                df = df.where(pd.notnull(df), None)

                for idx, row in df.iterrows():
                    try:
                        # Get head_id from head_code
                        cursor.execute(
                            "SELECT id FROM expenditure_heads WHERE head_code = ?",
                            (row.get('Head Code'),)
                        )
                        head = cursor.fetchone()
                        if head:
                            cursor.execute("""
                                INSERT OR REPLACE INTO expenditure_items (
                                    assignment_id, head_id, estimated_amount, actual_amount, remarks
                                ) VALUES (?, ?, ?, ?, ?)
                            """, (
                                row.get('Assignment ID'), head['id'],
                                row.get('Estimated (L)'), row.get('Actual (L)'),
                                row.get('Remarks')
                            ))
                            success_counts['expenditure'] += 1
                    except Exception as e:
                        errors.append(f"Expenditure row {idx + 2}: {str(e)}")

            # Import Config Options
            if 'Config Options' in xl.sheet_names:
                df = pd.read_excel(xl, 'Config Options')
                df = df.where(pd.notnull(df), None)

                for idx, row in df.iterrows():
                    try:
                        cursor.execute("""
                            INSERT OR REPLACE INTO config_options (
                                category, option_value, option_label, parent_value, sort_order, is_active
                            ) VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            row.get('Category'), row.get('Value'), row.get('Label'),
                            row.get('Parent Value'), row.get('Sort Order'),
                            1 if row.get('Active') in [1, '1', 'Yes', 'yes', True, None] else 0
                        ))
                        success_counts['config'] += 1
                    except Exception as e:
                        errors.append(f"Config row {idx + 2}: {str(e)}")

    except Exception as e:
        errors.append(f"File processing error: {str(e)}")

    return templates.TemplateResponse("data_import.html", {
        "request": request,
        "user": user,
        "success_counts": success_counts,
        "errors": errors[:20],  # Limit errors shown
        "total_errors": len(errors)
    })


@router.get("/admin/config", response_class=HTMLResponse)
async def config_page(request: Request):
    """Display configuration management page. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    with get_db() as conn:
        cursor = conn.cursor()

        # Get all config options grouped by category
        cursor.execute("""
            SELECT category, option_value, option_label, parent_value, sort_order, is_active
            FROM config_options
            ORDER BY category, sort_order
        """)
        all_options = [dict(row) for row in cursor.fetchall()]

        # Group by category
        config_by_category = {}
        for opt in all_options:
            cat = opt['category']
            if cat not in config_by_category:
                config_by_category[cat] = []
            config_by_category[cat].append(opt)

        # Get domains for sub-domain parent dropdown
        domains = [opt for opt in all_options if opt['category'] == 'domain']

    return templates.TemplateResponse("admin_config.html", {
        "request": request,
        "user": user,
        "config_by_category": config_by_category,
        "domains": domains
    })


@router.post("/admin/config/add")
async def add_config_option(request: Request):
    """Add a new configuration option. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    form = await request.form()
    category = form.get('category')
    option_value = form.get('option_value')
    option_label = form.get('option_label')
    parent_value = form.get('parent_value') or None
    sort_order = int(form.get('sort_order', 0))

    with get_db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO config_options (category, option_value, option_label, parent_value, sort_order)
                VALUES (?, ?, ?, ?, ?)
            """, (category, option_value, option_label, parent_value, sort_order))
        except Exception as e:
            pass  # Ignore duplicates

    return RedirectResponse(url="/data/admin/config", status_code=302)


@router.post("/admin/config/update")
async def update_config_option(request: Request):
    """Update a configuration option. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    form = await request.form()
    category = form.get('category')
    old_value = form.get('old_value')
    new_value = form.get('new_value')
    new_label = form.get('new_label')
    parent_value = form.get('parent_value') or None
    sort_order = int(form.get('sort_order', 0))
    is_active = 1 if form.get('is_active') else 0

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE config_options SET
                option_value = ?, option_label = ?, parent_value = ?,
                sort_order = ?, is_active = ?
            WHERE category = ? AND option_value = ?
        """, (new_value, new_label, parent_value, sort_order, is_active, category, old_value))

        # If domain value changed, update sub-domains parent_value
        if category == 'domain' and old_value != new_value:
            cursor.execute("""
                UPDATE config_options SET parent_value = ?
                WHERE category = 'sub_domain' AND parent_value = ?
            """, (new_value, old_value))

    return RedirectResponse(url="/data/admin/config", status_code=302)


@router.post("/admin/config/delete")
async def delete_config_option(request: Request):
    """Delete a configuration option. Admin only."""
    user, redirect = require_admin_access(request)
    if redirect:
        return redirect

    form = await request.form()
    category = form.get('category')
    option_value = form.get('option_value')

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM config_options WHERE category = ? AND option_value = ?
        """, (category, option_value))

        # If deleting a domain, also delete its sub-domains
        if category == 'domain':
            cursor.execute("""
                DELETE FROM config_options
                WHERE category = 'sub_domain' AND parent_value = ?
            """, (option_value,))

    return RedirectResponse(url="/data/admin/config", status_code=302)


# API endpoint for getting sub-domains based on domain
@router.get("/api/subdomains/{domain}")
async def get_subdomains(domain: str):
    """Get sub-domains for a given domain."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT option_value, option_label FROM config_options
            WHERE category = 'sub_domain' AND parent_value = ? AND is_active = 1
            ORDER BY sort_order
        """, (domain,))
        return [{"value": row['option_value'], "label": row['option_label']} for row in cursor.fetchall()]

