"""
Data export/import views for managing assignment data via Excel.
Admin-only access for data management.
"""
import io
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from core.models import (
    Assignment, Client, ConfigOption, ExpenditureHead, ExpenditureItem,
    Milestone, Officer, Office, RevenueShare,
)


def _require_admin(user):
    """Check if user is admin."""
    return user.is_staff or user.admin_role_id == "ADMIN"


@login_required
def export_page(request):
    """Display data export options page."""
    if not _require_admin(request.user):
        return redirect("core:dashboard")
    return render(request, "data_management/export.html")


@login_required
def export_assignments(request):
    """Export all assignments to Excel."""
    if not _require_admin(request.user):
        return redirect("core:dashboard")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Sheet 1: Assignments
    ws = wb.active
    ws.title = "Assignments"
    headers = [
        "ID", "Assignment No", "Type", "Title", "Client", "Client Type",
        "City", "Domain", "Sub-Domain", "Office", "Status",
        "Total Value (L)", "Invoice Raised (L)", "Payment Received (L)",
        "Physical Progress %", "Timeline Progress %",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, a in enumerate(
        Assignment.objects.order_by("office__office_id", "assignment_no"), 2
    ):
        values = [
            a.pk, a.assignment_no, a.type, a.title, a.client, a.client_type,
            a.city, a.domain, a.sub_domain, a.office_id, a.status,
            a.total_value, a.invoice_amount, a.amount_received,
            a.physical_progress_percent, a.timeline_progress_percent,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Sheet 2: Milestones
    ws2 = wb.create_sheet("Milestones")
    ms_headers = [
        "Assignment ID", "Milestone No", "Title", "Target Date",
        "Status", "Invoice Raised", "Payment Received",
    ]
    for col, header in enumerate(ms_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, m in enumerate(
        Milestone.objects.order_by("assignment_id", "milestone_no"), 2
    ):
        values = [
            m.assignment_id, m.milestone_no, m.title, str(m.target_date) if m.target_date else "",
            m.status, m.invoice_raised, m.payment_received,
        ]
        for col_idx, value in enumerate(values, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    # Sheet 3: Revenue Shares
    ws3 = wb.create_sheet("Revenue Shares")
    rs_headers = ["Assignment ID", "Officer ID", "Share %", "Share Amount (L)"]
    for col, header in enumerate(rs_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, rs in enumerate(
        RevenueShare.objects.order_by("assignment_id", "officer_id"), 2
    ):
        values = [rs.assignment_id, rs.officer_id, rs.share_percent, rs.share_amount]
        for col_idx, value in enumerate(values, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    # Sheet 4: Officers (Reference)
    ws4 = wb.create_sheet("Officers (Ref)")
    off_headers = ["Officer ID", "Name", "Designation", "Office", "Email", "Annual Target"]
    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    for col, header in enumerate(off_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = green_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, o in enumerate(
        Officer.objects.filter(is_active=True).order_by("office_id", "name"), 2
    ):
        values = [o.officer_id, o.name, o.designation, o.office_id, o.email, o.annual_target]
        for col_idx, value in enumerate(values, 1):
            ws4.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    # Sheet 5: Config Options
    ws5 = wb.create_sheet("Config Options")
    cfg_headers = ["Category", "Value", "Label", "Parent Value", "Sort Order", "Active"]
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    for col, header in enumerate(cfg_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = orange_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, c in enumerate(
        ConfigOption.objects.order_by("category", "sort_order"), 2
    ):
        values = [c.category, c.option_value, c.option_label, c.parent_value, c.sort_order, c.is_active]
        for col_idx, value in enumerate(values, 1):
            ws5.cell(row=row_idx, column=col_idx, value=value).border = thin_border

    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"PMS_Data_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@login_required
def import_page(request):
    """Display data import page."""
    if not _require_admin(request.user):
        return redirect("core:dashboard")
    return render(request, "data_management/import.html")


@login_required
def import_data(request):
    """Import data from Excel file."""
    if request.method != "POST":
        return redirect("core:import_page")
    if not _require_admin(request.user):
        return redirect("core:dashboard")

    import openpyxl

    errors = []
    success_counts = {
        "assignments": 0, "milestones": 0, "revenue_shares": 0,
        "expenditure": 0, "config": 0,
    }

    uploaded = request.FILES.get("file")
    if not uploaded:
        errors.append("No file uploaded")
        return render(request, "data_management/import.html", {
            "success_counts": success_counts, "errors": errors, "total_errors": 1,
        })

    try:
        wb = openpyxl.load_workbook(uploaded, read_only=True)

        # Import Config Options
        if "Config Options" in wb.sheetnames:
            ws = wb["Config Options"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for idx, row in enumerate(rows):
                try:
                    if row[0] and row[1]:
                        ConfigOption.objects.update_or_create(
                            category=str(row[0]).strip(),
                            option_value=str(row[1]).strip(),
                            defaults={
                                "option_label": str(row[2] or "").strip(),
                                "parent_value": str(row[3] or "").strip(),
                                "sort_order": int(row[4] or 0),
                                "is_active": bool(row[5]) if row[5] is not None else True,
                            },
                        )
                        success_counts["config"] += 1
                except Exception as e:
                    errors.append(f"Config row {idx + 2}: {e}")

        wb.close()
    except Exception as e:
        errors.append(f"File processing error: {e}")

    return render(request, "data_management/import.html", {
        "success_counts": success_counts,
        "errors": errors[:20],
        "total_errors": len(errors),
    })


@login_required
def config_page(request):
    """Display configuration management page."""
    if not _require_admin(request.user):
        return redirect("core:dashboard")

    all_options = ConfigOption.objects.order_by("category", "sort_order")
    config_by_category = {}
    for opt in all_options:
        config_by_category.setdefault(opt.category, []).append(opt)

    domains = [opt for opt in all_options if opt.category == "domain"]

    return render(request, "data_management/config.html", {
        "config_by_category": config_by_category,
        "domains": domains,
    })


@login_required
def add_config_option(request):
    """Add a new configuration option."""
    if request.method != "POST" or not _require_admin(request.user):
        return redirect("core:config_page")

    ConfigOption.objects.get_or_create(
        category=request.POST.get("category", ""),
        option_value=request.POST.get("option_value", ""),
        defaults={
            "option_label": request.POST.get("option_label", ""),
            "parent_value": request.POST.get("parent_value", ""),
            "sort_order": int(request.POST.get("sort_order", 0)),
        },
    )
    return redirect("core:config_page")


@login_required
def update_config_option(request):
    """Update a configuration option."""
    if request.method != "POST" or not _require_admin(request.user):
        return redirect("core:config_page")

    category = request.POST.get("category", "")
    old_value = request.POST.get("old_value", "")
    new_value = request.POST.get("new_value", "")

    try:
        opt = ConfigOption.objects.get(category=category, option_value=old_value)
        opt.option_value = new_value
        opt.option_label = request.POST.get("new_label", "")
        opt.parent_value = request.POST.get("parent_value", "")
        opt.sort_order = int(request.POST.get("sort_order", 0))
        opt.is_active = bool(request.POST.get("is_active"))
        opt.save()

        # If domain value changed, update sub-domain parent values
        if category == "domain" and old_value != new_value:
            ConfigOption.objects.filter(
                category="sub_domain", parent_value=old_value,
            ).update(parent_value=new_value)
    except ConfigOption.DoesNotExist:
        pass

    return redirect("core:config_page")


@login_required
def delete_config_option(request):
    """Delete a configuration option."""
    if request.method != "POST" or not _require_admin(request.user):
        return redirect("core:config_page")

    category = request.POST.get("category", "")
    option_value = request.POST.get("option_value", "")

    ConfigOption.objects.filter(category=category, option_value=option_value).delete()
    if category == "domain":
        ConfigOption.objects.filter(category="sub_domain", parent_value=option_value).delete()

    return redirect("core:config_page")


@login_required
def api_get_subdomains(request, domain):
    """Get sub-domains for a given domain."""
    from django.http import JsonResponse
    options = ConfigOption.objects.filter(
        category="sub_domain", parent_value=domain, is_active=True,
    ).order_by("sort_order").values("option_value", "option_label")
    return JsonResponse(
        [{"value": o["option_value"], "label": o["option_label"]} for o in options],
        safe=False,
    )
